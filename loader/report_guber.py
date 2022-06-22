import pandas as pd 
import openpyxl, shutil, os, datetime, glob
from openpyxl.utils.dataframe import dataframe_to_rows

from system import table_one_column
from base import covid_sql, covid_insert
from clas import Dir

header_vp = ['idRows','nameMO','indicators','vp_Received_Count_All_SPb',
        'vp_Received_Count_All_LO','vp_Received_Count_toDay_Spb',
        'vp_Received_Count_toDay_LO','vp_Discharged_Count_All_SPb'
        ,'vp_Discharged_Count_All_LO','vp_Discharged_Count_toDay_Spb',
        'vp_Discharged_Count_toDay_LO','vp_Died_Count_All_SPb',
        'vp_Died_Count_All_LO','vp_Died_Count_toDay_Spb','vp_Died_Count_toDay_LO'
        ,'vp_Hospital_Count_All','vp_Hospital_Count_Spb','vp_Hospital_Count_LO'
        ,'vp_Hospital_Count_Ivl'] 

header_cv = ['idRows','nameMO','indicators','cv_Diagnosis_Count_All_SPb',
        'cv_Diagnosis_Count_All_LO','cv_Diagnosis_Count_toDay_Spb',
        'cv_Diagnosis_Count_toDay_LO','cv_Discharged_Count_All_SPb'
        ,'cv_Discharged_Count_All_LO','cv_Discharged_Count_toDay_Spb',
        'cv_Discharged_Count_toDay_LO','cv_Died_Count_All_SPb',
        'cv_Died_Count_All_LO','cv_Died_Count_toDay_Spb','cv_Died_Count_toDay_LO'
        ,'cv_Hospital_Count_All','cv_Hospital_Count_Spb','cv_Hospital_Count_LO'
        ,'cv_Hospital_Count_Ivl']

header_ivl = ['idRows','nameMO','ivl_Invaz_Count_All','ivl_Invaz_Count_Busy'
        ,'ivl_Invaz_Count_Free_All','ivl_Invaz_Count_Faulty',
        'ivl_NeInvaz_Count_All','ivl_NeInvaz_Count_Busy','ivl_NeInvaz_Count_Free_All'
        ,'ivl_NeInvaz_Count_Faulty','ivl_Pacient_Count_All','ivl_Pacient_Count_Covid']

header_bunk = ['idRows','nameMO','bn_Count_All','bn_Count_Ill_All',
        'bn_Count_Ill_Faulty','bn_Count_Ill_Free']


def load_sheet(file, sheetName, ColumsName, startRows, header_): 
    df = pd.read_excel(file, sheet_name= sheetName,header = None , usecols=ColumsName,  skiprows = startRows)
    df = df.set_axis(header_, axis=1, inplace=False)
    df["idRows"] = pd.to_numeric(df["idRows"]) 
    df = df.sort_values(["idRows"])
    df = df.loc[df["idRows"].notnull()]
    df = df.drop_duplicates()
    df = df.fillna(0)
    return df

def load_file(file, sheetName, ColumsName, startRows, header_, tableName):
    try:
        excel = load_sheet(file, sheetName, ColumsName, startRows, header_)
    except:
        return 1

    covid_insert(excel, tableName, 'mon_vp', False, 'append')

def check_data_tab(name):
    sql=f"""
    IF (EXISTS (SELECT * FROM {name})) 
    SELECT 1 ELSE SELECT 0 """
    return covid_sql(sql).iat[0,0]

def UpdateShablonFile(workb, nameSheet, svod, startRows):
    ws = workb[nameSheet]
    rows = dataframe_to_rows(svod,index=False, header=False)
    for r_idx, row in enumerate(rows, startRows):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

class guber_09_debtors(Exception):
    pass

def report_guber():
    
    DIRECTORY =  Dir.get('MG')
    PATH = DIRECTORY + '/' + datetime.datetime.now().strftime("%Y%m%d")
    
    try:
        os.mkdir(PATH)
    except:
        pass

    # заливка файликов в базу
    for FILE in glob.glob(DIRECTORY + '/из_почты/[!~$]*.xls*'):

        load_file(FILE, 'Cвод_ОРВИ_и_Пневм', 'A:S', 4, header_vp,   'ReportGubernator_Pnevm')
        load_file(FILE, 'Свод_COVID'       , 'A:S', 4, header_cv,   'ReportGubernator_Covid')
        load_file(FILE, 'Свод_ИВЛ'         , 'A:L', 3, header_ivl,  'ReportGubernator_Ivl'  )
        load_file(FILE, 'Свод_Койки'       , 'A:F', 2, header_bunk, 'ReportGubernator_Bunk' )
        
        os.replace(FILE, PATH + '/' + os.path.basename(FILE))

    # проверка должников
    if check_data_tab('mon_vp.v_DebtorsReportGubernator'):
        df = covid_sql("SELECT [Наименование МО] as 'должники!' FROM mon_vp.v_DebtorsReportGubernator")
        TABLE = table_one_column(df)
        raise guber_09_debtors(TABLE)
    
    file1 = 'help/09_стационары_для_Справки_Губернатора.xlsx'
    file2 = 'help/09_стационары_для_Справки_Губернатора2.xlsx'
    
    new_file1 = 'temp/09_стационары для Справки Губернатора_'+ datetime.datetime.now().strftime("%d.%m.%Y_%H_%M") + '.xlsx'
    new_file2 = 'temp/09_стационары для Справки Губернатора_'+ datetime.datetime.now().strftime("%d.%m.%Y") + '.xlsx'

    shutil.copyfile(file1, new_file1)
    shutil.copyfile(file2, new_file2)
    
    df_covid = covid_sql('SELECT * FROM [mon_vp].[v_GrandReport_Guber_Covid] order by idRows')
    df1_covid = df_covid.drop('Установлены диагнозы: вчера',1)\
            .drop('Установлены диагнозы: должно быть',1)\
            .drop('Установлены диагнозы: фактически',1)\
            .drop('на стационарном лечении: вчерашние данные',1)\
            .drop('на стационарном лечении: должно быть',1)\
            .drop('на стационарном лечении: фактически',1)

    df_pnev = covid_sql('SELECT * FROM [mon_vp].[v_GrandReport_Guber_Pnev] order by idRows')
    df1_pnev = df_pnev.drop('Установлены диагнозы: вчера',1)\
            .drop('Установлены диагнозы: должно быть',1)\
            .drop('Установлены диагнозы: фактически',1)\
            .drop('на стационарном лечении: вчерашние данные',1)\
            .drop('на стационарном лечении: должно быть',1)\
            .drop('на стационарном лечении: фактически',1)

    df_ivl = covid_sql('SELECT * FROM [mon_vp].[v_GrandReport_Guber_Ivl] order by idRows')

    df_bunk = covid_sql('SELECT * FROM [mon_vp].[v_GrandReport_Guber_Bunk] order by idRows')

    df_sys = covid_sql('SELECT * FROM mon_vp.v_GrandReport')

    df1_sys = df_sys.loc[df_sys.typeMO==1].sort_values(["numSort"]).drop('typeMO',1).drop('numSort',1)
    df2_sys = df_sys.loc[df_sys.typeMO==2].sort_values(["numSort"]).drop('typeMO',1).drop('numSort',1)

    wb = openpyxl.load_workbook(new_file1)

    UpdateShablonFile(wb, 'Cвод_ОРВИ_и_Пневм', df1_pnev, 6)
    UpdateShablonFile(wb, 'Свод_COVID', df1_covid, 6)
    UpdateShablonFile(wb, 'Свод_ИВЛ', df_ivl, 5)
    UpdateShablonFile(wb, 'Свод_Койки', df_bunk, 4)

    if not check_data_tab('mon_vp.v_DebtorsReport'):
        UpdateShablonFile(wb, 'Отчет_СЮС', df1_sys, 9)
        UpdateShablonFile(wb, 'Отчет_СЮС', df2_sys, 73)

    wb.save(new_file1)

    wb1 = openpyxl.load_workbook(new_file2)

    UpdateShablonFile(wb1, 'Свод', df_pnev, 8)
    UpdateShablonFile(wb1, 'Свод', df_covid, 97)

    ws = wb1['Свод']
    ws['Q2'] = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%d.%m.%Y')

    wb1.save(new_file2)

    shutil.copyfile(new_file1, DIRECTORY + '/' + new_file1[5:] )
    shutil.copyfile(new_file2, DIRECTORY + '/' + new_file2[5:] )

    return new_file1 + ';' + new_file2
