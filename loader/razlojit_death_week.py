from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
import shutil

from base import covid_sql, covid_insert
from clas import Dir


COLUMNS = [
    'Район проживания', 'Дата начала заболевания',
    'Факт обращения за медицинской помощью на амбулаторном этапе (да/нет)',
    'Дата обращения за медицинской помощью  на амбулаторном этапе',
    'Факт выполнения КТ на амбулаторном этапе (да/нет)',
    'Факт выполнения ПЦР-SARS-CoV-2  на амбулаторном этапе (да/нет)',
    'Факт получения бесплатной лекарственной терапии (БЛТ) на амбулаторном этапе (да/нет)',
    'Дата госпитализации',
    'Степень тяжести состояния при госпитализации (легкая, ср.тяжести, тяжелая)',
    'Поступление в ОРИТ  при госпитализации (да/нет)',
    'Смерть наступила в первые сутки с момента госпитализации (да/нет)',
    'Факт получения антицитокиновой терапии в стационаре (да/нет)'
    ]


def razlojit_death_week():
    """Раскладываем по FTP папкам умерших за неделю,
    согласно федеральному регистру ковида,
    чтобы организации заполнили доп информацию по случаем смерти"""

    DATE_END = (datetime.today() + relativedelta(weeks=-1, weekday=2)).date()
    DATE_START = DATE_END - timedelta(days=6)

    SQL = f"""
        select
            dbo.get_Gid(idPatient) as 'Gid',
            [Медицинская организация], [ФИО], [Дата рождения]
            ,dbo.[f_calculation_age]([Дата рождения], [Дата исхода заболевания]) as 'Возраст'
            ,[Посмертный диагноз]
        from robo.v_FedReg
        where [Исход заболевания] = 'Смерть'
            and [Дата исхода заболевания]  BETWEEN '{DATE_START}' AND '{DATE_END}'
            -- and YEAR([Дата исхода заболевания]) = YEAR(getdate())
            and ([Посмертный диагноз]  in ('U07.1','U07.2')
                or [Посмертный диагноз] like 'J1[2-8]%' )
            and [Субъект РФ] = 'г. Санкт-Петербург'
        """

    DF = covid_sql(SQL)

    for COL in COLUMNS:
        DF[COL] = ''

    SQL = 'SELECT [Наименование в ФР], [user] from robo.directory'
    MO_DIR = covid_sql(SQL)

    STAT = pd.DataFrame()
    for MO in DF['Медицинская организация'].unique():
        k = len(STAT)
        STAT.loc[k, 'Медицинская организация'] = MO

        MO_FR = MO.replace(' (стац)', '').replace(' (амб.)', '')
        try:
            MO_USER = MO_DIR.loc[MO_DIR['Наименование в ФР'] == MO_FR, 'user'][0]
        except KeyError:
            STAT.loc[k, 'комментарий'] = 'не найдена МО в справочнике'
            continue

        DIR = Dir.get('covid') + MO_USER + 'Умершие за неделю'

        try:
            os.mkdir(DIR)
        except FileExistsError:
            pass

        OTCHET = DF[DF['Медицинская организация'] == MO]
        NAME_PART = {
            '(стац)' in MO: '/умершие (стац) с ',
            '(амб.)' in MO: '/умершие (амб.) с ',
            }.get([True], '/умершие с ')

        FILE_NAME = DIR \
            + NAME_PART \
            + f'{DATE_START} по {DATE_END}.xlsx'

        shutil.copyfile('help/death_week_shablon.xlsx', FILE_NAME)

        wb = load_workbook(FILE_NAME)

        ws = wb['death_week']

        rows = dataframe_to_rows(OTCHET, index=False, header=False)
        for r_idx, row in enumerate(rows, 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(FILE_NAME)
        STAT.loc[k, 'комментарий'] = FILE_NAME

    STAT_FILE = 'temp/разложенные файлы по умершим за неделю.xlsx'
    with pd.ExcelWriter(STAT_FILE) as writer:
        STAT.to_excel(writer, sheet_name='файлы', index=False)
        DF.to_excel(writer, sheet_name='свод', index=False)

    REPORT = pd.DataFrame(DF['Gid'])

    REPORT['time'] = datetime.now()
    REPORT['Dates'] = f'{DATE_START} - {DATE_END}'
    covid_insert(REPORT, 'report_deth_week', 'robo', False, 'append')

    return STAT_FILE
