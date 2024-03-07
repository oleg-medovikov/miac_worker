import pandas as pd
import openpyxl, shutil, os, datetime, glob
from openpyxl.utils.dataframe import dataframe_to_rows

from system import table_one_column
from base import covid_sql, covid_insert
from clas import Dir


def load_file_mo(FILE):
    nameMO = pd.read_excel(
        FILE, sheet_name="Титул", header=3, usecols="H", nrows=1
    ).iloc[0, 0]

    df = pd.read_excel(FILE, sheet_name="Данные1", header=6, usecols="C:AH", nrows=1)

    df = df.fillna(0)
    #        df['nameMO'] = nameMO
    return df, nameMO


def check_data_table(name):
    sql = f"""
    IF (EXISTS (SELECT * FROM {name})) 
        SELECT 1 ELSE SELECT 0 """

    return covid_sql(sql).iat[0, 0]


class my_except(Exception):
    pass


def report_vp_and_cv():
    FILES = glob.glob(Dir.get("VP_CV") + "/из_почты/[!~$]*.xlsx")

    # if len(FILES) == 0:
    #    raise my_except("Папка пустая!")

    PATH = Dir.get("VP_CV") + "/" + datetime.datetime.now().strftime("%Y%m%d")

    if not os.path.exists(PATH):
        try:
            os.mkdir(PATH)
        except OSError:
            raise my_except("не смог создать папку")

    ERROR = pd.DataFrame()
    list_ = []

    for FILE in FILES:
        try:
            EXCEL, nameMO = load_file_mo(FILE)
            if len(EXCEL.dtypes.unique()) > 1:
                raise my_except(f"В файле {FILE.rsplit('/',1)[1]} неправильные данные")
            else:
                EXCEL["nameMO"] = nameMO
        except Exception as e:
            # raise my_except(str(e))
            ERROR.loc[len(ERROR), "Не обработаны файлы"] = FILE.rsplit("/", 1)[1]
        else:
            list_.append(EXCEL)

    if len(ERROR):
        STRING = table_one_column(ERROR)
        raise my_except(STRING)

    for FILE in FILES:
        os.replace(FILE, PATH + "/" + os.path.basename(FILE))

    if len(list_):
        df = pd.concat(list_)

        HEADER = [
            "vp_03_Power_Count_Departments",
            "vp_04_Power_Count_Allocated_All",
            "vp_05_Power_Count_Allocated",
            "vp_06_Power_Count_Reanimation_All",
            "vp_07_Power_Count_Reanimation",
            "vp_08_Hospital_All",
            "vp_09_Hospital_Day",
            "vp_10_Hospital_Hard_All",
            "vp_11_Hospital_Hard_Reaniamation",
            "vp_12_Hospital_Hard_Ivl",
            "vp_13_ReceivedAll",
            "vp_14_ReceivedDay",
            "vp_15_DischargedAll",
            "vp_16_DischargedDay",
            "vp_17_DiedAll",
            "vp_18_DiedDay",
            "cv_19_Power_Count_Departments",
            "cv_20_Power_Count_Allocated_All",
            "cv_21_Power_Count_Allocated",
            "cv_22_Power_Count_Reanimation_All",
            "cv_23_Power_Count_Reanimation",
            "cv_24_Hospital_All",
            "cv_25_Hospital_Day",
            "cv_26_Hospital_Hard_All",
            "cv_27_Hospital_Hard_Reaniamation",
            "cv_28_Hospital_Hard_Ivl",
            "cv_29_ReceivedAll",
            "cv_30_ReceivedDay",
            "cv_31_DischargedAll",
            "cv_32_DischargedDay",
            "cv_33_DiedAll",
            "cv_34_DiedDay",
            "nameMO",
        ]

        df.set_axis(HEADER, axis=1, inplace=True)

        covid_insert(df, "HistoryFileMO", "mon_vp", False, "append")

    if check_data_table("mon_vp.v_DebtorsReport"):
        TABLE = covid_sql(
            """SELECT [Наименование МО] as 'должники!'
                                FROM mon_vp.v_DebtorsReport"""
        )
        STRING = table_one_column(TABLE)
        raise my_except(STRING)

    # === данные загружены в базу, приступаем к созданию отчёта
    # === выгружаем исходные данные в первую вкладку отчёта

    df = covid_sql("SELECT * FROM mon_vp.v_GrandReport")

    df1 = (
        df.loc[df.typeMO == 1]
        .sort_values(["numSort"])
        .drop("typeMO", 1)
        .drop("numSort", 1)
    )

    df2 = (
        df.loc[df.typeMO == 2]
        .sort_values(["numSort"])
        .drop("typeMO", 1)
        .drop("numSort", 1)
    )

    DATE = datetime.datetime.now() + datetime.timedelta(days=1)

    NEW_FILE = "temp/СводОбщий_" + DATE.strftime("%d %m %Y") + ".xlsx"

    shutil.copyfile("help/шаблон Мониторинг ВП.xlsx", NEW_FILE)

    wb = openpyxl.load_workbook(NEW_FILE)
    ws = wb["Данные"]
    ws.cell(row=1, column=2, value=DATE.strftime("%d.%m.%Y"))

    rows = dataframe_to_rows(df1, index=False, header=False)
    for r_idx, row in enumerate(rows, 9):
        for c_idx, value in enumerate(row, 1):
            try:
                ws.cell(row=r_idx, column=c_idx, value=value)
            except:
                continue
    rows = dataframe_to_rows(df2, index=False, header=False)
    for r_idx, row in enumerate(rows, 73):
        for c_idx, value in enumerate(row, 1):
            try:
                ws.cell(row=r_idx, column=c_idx, value=value)
            except:
                continue
    wb.save(NEW_FILE)

    # == сейчас мы выгрузим и сохраним данные по проверкам

    df = covid_sql("exec mon_vp.p_CheckMonitorVpAndCovid")

    part_one = df.iloc[:, range(26)]
    part_two = df.iloc[:, [0] + list(range(26, 58, 1))]

    wb = openpyxl.load_workbook(NEW_FILE)
    ws = wb["Проверка"]

    rows = dataframe_to_rows(part_one, index=False, header=False)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, value in enumerate(row, 1):
            try:
                ws.cell(row=r_idx, column=c_idx, value=value)
            except:
                continue
    ws = wb["Разница"]

    rows = dataframe_to_rows(part_two, index=False, header=False)
    for r_idx, row in enumerate(rows, 7):
        for c_idx, value in enumerate(row, 1):
            try:
                ws.cell(row=r_idx, column=c_idx, value=value)
            except:
                continue
    wb.save(NEW_FILE)

    try:
        shutil.copyfile(NEW_FILE, Dir.get("VP_CV") + "/" + NEW_FILE[5:])
    except PermissionError:
        raise my_except("Закройте файлик! Не могу скопировать")

    return NEW_FILE
