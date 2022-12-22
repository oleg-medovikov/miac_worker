import glob
from pandas import DataFrame
from openpyxl import load_workbook

from clas import Dir
from .ColumnName import NAMES, ColumnName


def check_header(FILE: str) -> str:
    "Проверка заголовка в первой строке файла"
    wb = load_workbook(FILE, data_only=True)
    ws = wb.active
    STR = ''
    COLUMNS = ColumnName.all_names(NAMES)
    COLUMNS.remove('Пол')

    for NAME in NAMES:
        for cell in ws[1]:
            if cell.value is None:
                continue

            if all(map(lambda x: x in cell.value.lower(), NAME.keys)):
                "если в ячейке найдены все ключи"
                if cell.value != NAME.name:

                    STR += f'\n"{cell.value}" исправлена на "{NAME.name}"'
                    cell.value = NAME.name
                    try:
                        COLUMNS.remove(NAME.name)
                    except KeyError:
                        pass
                else:
                    try:
                        COLUMNS.remove(NAME.name)
                    except KeyError:
                        pass
                    continue

    for COL in COLUMNS:
        STR += f'\nне найдена колонка "{COL}"'

    if STR == '':
        STR = 'Файл соответствует шаблону'

    wb.save(FILE)

    return STR


def check_files_cardio():
    "Функция проверяет шапки внутри файлов кардио"

    PATH = Dir.get('CARDIO')
    MASK = PATH + '/ori.cardio.[!1]*/*_122/[!~]*.xls*'
    DF = DataFrame()

    for FILE in glob.iglob(MASK):
        k = len(DF)
        DF.loc[k, 'file'] = FILE
        try:
            MESS = check_header(FILE)
        except Exception as e:
            MESS = 'Файл не читается \n' + str(e)

        MESS_FILE = FILE[:-5] + ' соответствие шаблону.txt'

        with open(MESS_FILE, 'w') as file:
            try:
                file.write(MESS)
            except PermissionError:
                pass

        DF.loc[k, 'mess'] = MESS

    OTCHET = 'temp/cardio_files_rename.xlsx'
    DF.to_excel(OTCHET)

    return OTCHET
