from datetime import datetime, timedelta
import shutil
import openpyxl
import requests
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from base import parus_sql, nsi_sql
from conf import URL_870


def svod_27_covid_19_regiz():
    DATA = requests.get(URL_870, verify=False).json()

    REGIZ = pd.DataFrame(DATA)
    columns = [
        'orderresponse_assign_organization_level1_key', 'ShortNameMO',
        'Кол-во тестов', 'Кол-во ПЦР тестов',
        'Кол-во положительных ПЦР тестов', 'Кол-во тестов на антитела',
        'Кол-во положительных тестов на антитела',
        'Кол-во тестов на антитела после вакцинации',
        'Кол-во положительных тестов на антитела после вакцинации'
        ]
    REGIZ = REGIZ[columns]

    SQL_1 = open('parus/sql/covid_27_regiz.sql', 'r').read()
    SQL_2 = open('parus/sql/nsi_27.sql', 'r').read()

    DF = parus_sql(SQL_1)
    NSI = nsi_sql(SQL_2)

    DATE = (datetime.now() - timedelta(days=2)).strftime('%d_%m_%Y')

    NEW_NAME = 'temp/' + DATE + '_27_COVID_19_regiz.xlsx'

    shutil.copyfile('help/27_COVID_19_regiz.xlsx', NEW_NAME)

    wb = openpyxl.load_workbook(NEW_NAME)

    ws = wb['parus']
    rows = dataframe_to_rows(DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['regiz']
    rows = dataframe_to_rows(REGIZ, index=False, header=True)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['nsi']
    rows = dataframe_to_rows(NSI, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME)

    return NEW_NAME
