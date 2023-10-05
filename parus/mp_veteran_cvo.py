import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql
from pandas import read_excel, DataFrame


def mp_veteran_cvo():
    SQL = open('parus/sql/mp_veteran_cvo.sql', 'r').read()
    SQL_2 = open('parus/sql/mp_veteran_cvo_itogo.sql', 'r').read()

    DF = parus_sql(SQL)
    SUM = parus_sql(SQL_2)
    ITOGO = DataFrame()

    for i in SUM.index:
        pok = SUM.at[i, 'POKAZATEL']
        index = int(pok.rsplit('_', 1)[-1])
        column = pok.rsplit('_', 2)[-2]

        ITOGO.loc[index, column] = SUM.at[i, 'VALUE']

    ITOGO = ITOGO.sort_index()

    DATE = DF.at[0, 'DAY']

    del DF['DAY']

    DF.loc['Итого по городу'] = DF.loc[
            DF['POK01'] != 'Итого по организации'
            ].sum(numeric_only=True, axis=0)
    DF.loc['Итого по городу', 'ORGANIZATION'] = 'Итого:'

    DOLG = read_excel('help/mp_veteran_cvo.xlsx', sheet_name='долг')
    for _ in DOLG.index:
        if DOLG.at[_, 'Список организаций'] in DF['ORGANIZATION'].unique():
            DOLG.loc[_, 'Наличие в отчете'] = 1
        else:
            DOLG.loc[_, 'Наличие в отчете'] = 0

    NEW_NAME = 'temp/МП_ветеранам_СВО_' + DATE + '.xlsx'

    shutil.copyfile('help/mp_veteran_cvo.xlsx', NEW_NAME)

    wb = openpyxl.load_workbook(NEW_NAME)
    ws = wb['Итого']
    rows = dataframe_to_rows(ITOGO, index=False, header=False)
    for r_idx, row in enumerate(rows, 5):
        for c_idx, value in enumerate(row, 3):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Свод']
    rows = dataframe_to_rows(DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 5):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['долг']
    rows = dataframe_to_rows(DOLG, index=False, header=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME)

    return NEW_NAME
