import time, datetime, shutil, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import pandas as pd
from base import parus_sql

POKS = ['_01', '_02', '_03', '_04', '_05', '_06', '_07', '_08', '_09',
        '_10', '_11', '_12', '_13', '_14', '_15', '_16', '_17', '_18',
        '_19', '_20', '_21', '_22', '_23', '_24', '_25', '_26', '_27',
        '_28', '_29', '_30', '_31', '_32', '_33', '_34', '_35', '_36',
        '_37', '_38', '_39', '_40', '_41', '_42', '_43', '_44', '_45',
        '_46', '_47', '_48', '_49', '_50', '_51', '_52', '_53', '_54',
        '_55', '_56', '_57', '_58', '_59']


def svod_4_3_covid_19():
    MEDICATIONS = open('parus/sql/medications', 'r').read().split('\n')

    NEW_NAME = 'temp/4.3 COVID 19.xlsx'

    shutil.copyfile('help/4.3_COVID_19_svod.xlsx', NEW_NAME )

    wb= openpyxl.load_workbook(NEW_NAME,data_only=False)

    for MED, POK in zip(MEDICATIONS, POKS):
        POKAZATELI = f"""
'4.3_1{POK}'pok1,
'4.3_2{POK}' pok2,
'4.3_3{POK}' pok3,
'4.3_4{POK}' pok4,
'4.3_5{POK}' pok5,
'4.3_6{POK}' pok6,
'4.3_7{POK}' pok7,
'4.3_8{POK}' pok8"""

        SQL = open('parus/sql/covid_4.3_svod_stac.sql', 'r')\
                .read().replace('pokazateli', POKAZATELI)
        DF = parus_sql( SQL )

        del DF ['DAY']

        DF['POK9'] = 30*(DF['POK1'] + DF['POK3']) / DF['POK4'] 
        DF['POK10'] =   (DF['POK1'] + DF['POK3']) / DF['POK4'] 
        
        ws = wb[POK.replace('_','')]

        # пишем наименование лекарства
        ws.cell(row=3, column=6, value=MED)
        # пишем значения в таблицу
        rows = dataframe_to_rows(DF,index=False, header=False)
        for r_idx, row in enumerate(rows,5):  
            for c_idx, value in enumerate(row, 5):
                ws.cell(row=r_idx, column=c_idx, value=value)

        
        ws = wb['main']
        ws.cell(row= 6  + MEDICATIONS.index(MED),
                column=5,
                value=MED)
        
        ws.cell(row= 6  + MEDICATIONS.index(MED),
                column=6, value=DF.loc[~(DF['POK1'].isnull()),
                'POK1'].sum())
        ws.cell(row= 6  + MEDICATIONS.index(MED),
                column=7, value=DF.loc[~(DF['POK3'].isnull()),
                'POK3'].sum())
        ws.cell(row= 6  + MEDICATIONS.index(MED),
                column=8, value=DF.loc[~(DF['POK4'].isnull()),
                'POK4'].sum())
        ws.cell(row= 6  + MEDICATIONS.index(MED),
                column=9, value=DF.loc[~(DF['POK5'].isnull()),
                'POK5'].sum())


        ws.cell(row= 6  + MEDICATIONS.index(MED),
                column=10, value=DF.loc[~(DF['POK9'].isnull()) & (DF['POK9'] != np.inf),
                'POK9'].mean())
        ws.cell(row= 6  + MEDICATIONS.index(MED),
                column=11,
                value=DF.loc[~(DF['POK10'].isnull()) & (DF['POK10'] != np.inf),
                'POK10'].mean())
        
        if len(DF.loc[(DF['POK9'].isnull()) | (DF['POK9'] == np.inf) , 'POK9']):
            ws.cell(row= 6  + MEDICATIONS.index(MED), column=12, value='Да')
        else:
            ws.cell(row= 6  + MEDICATIONS.index(MED), column=12, value='Нет')


    wb.save( NEW_NAME )

    return NEW_NAME
