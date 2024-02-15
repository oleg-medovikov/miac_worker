from datetime import datetime
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql
from system import return_weekday


def grip_and_ori():
    DAY = return_weekday(datetime.now(), 2)

    SQL_A = open("parus/sql/grip_and_ori_A.sql", "r").read()
    SQL_B = open("parus/sql/grip_and_ori_B.sql", "r").read()
    SQL_D = open("parus/sql/grip_and_ori_D.sql", "r").read()
    SQL_G = open("parus/sql/grip_and_ori_G.sql", "r").read()

    SQL_A = SQL_A.replace("__DAY__", DAY)
    SQL_B = SQL_B.replace("__DAY__", DAY)
    SQL_D = SQL_D.replace("__DAY__", DAY)
    SQL_G = SQL_G.replace("__DAY__", DAY)

    DF_A = parus_sql(SQL_A)
    DF_B = parus_sql(SQL_B)
    DF_D = parus_sql(SQL_D)
    DF_G = parus_sql(SQL_G)

    DF_A = DF_A.pivot_table(
        index=["ORGANIZATION"], columns=["POKAZATEL"], aggfunc="first"
    ).stack(0)

    DF_B = DF_B.pivot_table(
        index=["ORGANIZATION"], columns=["POKAZATEL"], aggfunc="first"
    ).stack(0)

    DF_D = DF_D.pivot_table(
        index=["ORGANIZATION"], columns=["POKAZATEL"], aggfunc="first"
    ).stack(0)

    DF_G = DF_G.pivot_table(
        index=["ORGANIZATION"], columns=["POKAZATEL"], aggfunc="first"
    ).stack(0)

    DF_A = DF_A.reset_index()
    DF_B = DF_B.reset_index()
    DF_D = DF_D.reset_index()
    DF_G = DF_G.reset_index()

    del DF_A["level_1"]
    del DF_B["level_1"]
    del DF_D["level_1"]
    del DF_G["level_1"]

    DF_A = DF_A[
        [
            "ORGANIZATION",
            "table_sum_01",
            "table_01",
            "table_02",
            "table_03",
            "table_04",
            "table_05",
        ]
    ]
    DF_B = DF_B[
        [
            "ORGANIZATION",
            "table_sum_02",
            "table_sum_03",
            "table_06",
            "table_07",
            "table_08",
            "table_09",
            "table_10",
            "table_11",
            "table_12",
            "table_13",
            "table_14",
            "table_15",
            "table_16",
            "table_17",
            "table_18",
            "table_19",
            "table_20",
            "table_21",
            "table_22",
            "table_23",
            "table_24",
            "table_25",
            "table_26",
            "table_27",
            "table_28",
            "table_29",
            "table_30",
            "table_31",
            "table_31.1",
            "table_31.2",
            "table_32",
            "table_33",
        ]
    ]
    DF_D = DF_D[
        [
            "ORGANIZATION",
            "table_34",
            "table_35",
            "table_36",
            "table_37",
            "table_38",
            "table_39",
        ]
    ]
    DF_G1 = DF_G[
        [
            "ORGANIZATION",
            "table_43",
            "table_44",
            "table_53",
            "table_54",
        ]
    ]
    DF_G2 = DF_G[
        [
            "ORGANIZATION",
            "table_55",
            "table_56",
            "table_57",
            "table_58",
            "table_59",
            "table_60",
            "table_61",
            "table_62",
            "table_45",
            "table_46",
            "table_47",
            "table_48",
            "table_49",
            "table_50",
            "table_51",
            "table_52",
        ]
    ]

    DATE = DAY.replace(".", "_")

    NEW_NAME_A = f"/tmp/О_сроках_лаб_исл_на_ОРИ_{DATE}.xlsx"
    NEW_NAME_B = f"/tmp/О_результатах_тест_ОРИ_{DATE}.xlsx"
    NEW_NAME_D = f"/tmp/О_наличии_тест_ОРИ_{DATE}.xlsx"
    NEW_NAME_G1 = f"/tmp/Запас_диагностических_тест-систем_{DATE}.xlsx"
    NEW_NAME_G2 = f"/tmp/Запас_диагностических_тест-систем_в_МО_{DATE}.xlsx"

    shutil.copyfile("help/grip_and_ori_a.xlsx", NEW_NAME_A)
    shutil.copyfile("help/grip_and_ori_b.xlsx", NEW_NAME_B)
    shutil.copyfile("help/grip_and_ori_d.xlsx", NEW_NAME_D)
    shutil.copyfile("help/grip_and_ori_g1.xlsx", NEW_NAME_G1)
    shutil.copyfile("help/grip_and_ori_g2.xlsx", NEW_NAME_G2)

    wb = load_workbook(NEW_NAME_A)

    ws = wb["svod"]
    rows = dataframe_to_rows(DF_A, index=False, header=False)
    for r_idx, row in enumerate(rows, 8):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME_A)

    wb = load_workbook(NEW_NAME_B)

    ws = wb["svod"]
    rows = dataframe_to_rows(DF_B, index=False, header=False)
    for r_idx, row in enumerate(rows, 10):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME_B)

    wb = load_workbook(NEW_NAME_D)

    ws = wb["svod"]
    rows = dataframe_to_rows(DF_D, index=False, header=False)
    for r_idx, row in enumerate(rows, 7):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME_D)

    wb = load_workbook(NEW_NAME_G1)

    ws = wb["Лист1"]
    rows = dataframe_to_rows(DF_G1, index=False, header=False)
    for r_idx, row in enumerate(rows, 7):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME_G1)

    wb = load_workbook(NEW_NAME_G2)

    ws = wb["Лист1"]
    rows = dataframe_to_rows(DF_G2, index=False, header=False)
    for r_idx, row in enumerate(rows, 9):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME_G2)

    return (
        NEW_NAME_B
        + ";"
        + NEW_NAME_A
        + ";"
        + NEW_NAME_D
        + ";"
        + NEW_NAME_G1
        + ";"
        + NEW_NAME_G2
    )
