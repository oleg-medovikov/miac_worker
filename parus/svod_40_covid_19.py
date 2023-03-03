from datetime import datetime, timedelta
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql

FILES = [
    'parus/sql/covid_40_exit.sql',
    'parus/sql/covid_40_spytnic.sql',
    'parus/sql/covid_40_spytnic_old.sql',
    'parus/sql/covid_40_epivak.sql',
    'parus/sql/covid_40_epivak_old.sql',
    'parus/sql/covid_40_covivak.sql',
    'parus/sql/covid_40_covivak_old.sql',
    'parus/sql/covid_40_revac.sql',
    'parus/sql/covid_40_revac_old_new.sql',
    'parus/sql/covid_40_light.sql',
    'parus/sql/covid_40_light_old.sql'
        ]


def change_sql(FILE: str, MO):
    """Меняем запрос, добавляя в него выбывшие организации"""
    sql = open(FILE, 'r').read()
    if len(MO) == 0:
        return sql

    MO_str = ''

    for _ in MO['ORG'].unique():
        MO_str += "'" + _ + "',"

    MO_str = MO_str[:-1]

    if 'old' in FILE:
        string = f"\t\t\tand ( ( r.BDATE = trunc(SYSDATE) - 2  and a.AGNNAME not in ( {MO_str} ) )"
        for i in range(len(MO)):
            string += f"\n\t\t\tOR(r.BDATE = TO_DATE('{MO.at[i,'DAY']}','DD.MM.YYYY') - 1  and a.AGNNAME = '{MO.at[i,'ORG']}' ) "
        string += ')'
    else:
        string = f"\t\t\tand (( r.BDATE = trunc(SYSDATE) - 1 and a.AGNNAME not in ({MO_str} ) )"
        for i in range(len(MO)):
            string += f"\n\t\t\tOR(r.BDATE = TO_DATE('{MO.at[i,'DAY']}','DD.MM.YYYY') and a.AGNNAME = '{MO.at[i,'ORG']}' ) "
        string += ')'

    for line in sql.split('\n'):
        if 'trunc' in line:
            sql = sql.replace(line, string)

    with open('temp/' + FILE.rsplit('/')[-1], 'w') as f:
        f.write(sql)

    return sql


class my_except(Exception):
    pass


def svod_40_covid_19():
    sql = open(FILES[0], 'r').read()
    MO = parus_sql(sql)

    sput = parus_sql(change_sql(FILES[1], MO))
    sput_old = parus_sql(change_sql(FILES[2], MO))

    epivak = parus_sql(change_sql(FILES[3], MO))
    epivak_old = parus_sql(change_sql(FILES[4], MO))

    covivak = parus_sql(change_sql(FILES[5], MO))
    covivak_old = parus_sql(change_sql(FILES[6], MO))

    revac = parus_sql(change_sql(FILES[7], MO))
    revac_old = parus_sql(change_sql(FILES[8], MO))

    light = parus_sql(change_sql(FILES[9], MO))
    light_old = parus_sql(change_sql(FILES[10], MO))

    """
    sput        = sput.loc[~sput[sput.columns[5]].isnull()]
    sput_old    = sput_old.loc[~sput_old[sput_old.columns[5]].isnull()]
    epivak      = epivak.loc[~epivak[epivak.columns[5]].isnull()]
    epivak_old  = epivak_old.loc[~epivak_old[epivak_old.columns[5]].isnull()]
    covivak     = covivak.loc[~covivak[covivak.columns[5]].isnull()]
    covivak_old = covivak_old.loc[~covivak_old[covivak_old.columns[5]].isnull()]
    revac       = revac.loc[~revac[revac.columns[6]].isnull()]
    revac_old   = revac_old.loc[~revac_old[revac_old.columns[6]].isnull()]
    light       = light.loc[~light[light.columns[5]].isnull()]
    light_old   = light_old.loc[~light_old[light_old.columns[5]].isnull()]
    """

    del sput['ORGANIZATION']
    del sput_old['ORGANIZATION']
    del epivak['ORGANIZATION']
    del epivak_old['ORGANIZATION']
    del covivak['ORGANIZATION']
    del covivak_old['ORGANIZATION']
    # del revac ['INDX']
    # del revac_old ['INDX']
    del light['ORGANIZATION']
    del light_old['ORGANIZATION']

    """
    sput.drop_duplicates(keep='first', inplace=True)
    sput_old.drop_duplicates(keep='first', inplace=True)
    epivak.drop_duplicates(keep='first', inplace=True)
    epivak_old.drop_duplicates(keep='first', inplace=True)
    covivak.drop_duplicates(keep='first', inplace=True)
    covivak_old.drop_duplicates(keep='first', inplace=True)
    revac.drop_duplicates(keep='first', inplace=True)
    revac_old.drop_duplicates(keep='first', inplace=True)
    light.drop_duplicates(keep='first', inplace=True)
    light_old.drop_duplicates(keep='first', inplace=True)
    """
    revac = revac.loc[revac['TIP'] == 'Медицинская организация']
    revac_old = revac_old.loc[revac_old['TIP'] == 'Медицинская организация']

    date_otch = (datetime.now() - timedelta(days=1)).strftime('%d.%m.%Y')

    new_name_pred = 'temp/40_COVID_19_БОТКИНА_' + date_otch + '_предварительный.xlsx'
    new_name_osn = 'temp/40_COVID_19_БОТКИНА_' + date_otch + '_основной.xlsx'

    shutil.copyfile('help/40_COVID_19_pred.xlsx', new_name_pred)
    shutil.copyfile('help/40_COVID_19_osn.xlsx', new_name_osn)

    # Записываем данные в предварительный файл
    wb = openpyxl.load_workbook(new_name_pred)

    DICT = {
        'Спутник-V':          (sput,        5, 1),
        'Вчера_Спутник':      (sput_old,    5, 1),
        'ЭпиВакКорона':       (epivak,      5, 1),
        'Вчера_ЭпиВак':       (epivak_old,  5, 1),
        'КовиВак':            (covivak,     5, 1),
        'Спутник Лайт':       (light,       5, 1),
        'Вчера_КовиВак':      (covivak_old, 5, 1),
        'Ревакцинация':       (revac,       9, 1),
        'Вчера_ревакцин':     (revac_old,   9, 1),
        'Вчера_Спутник Лайт': (light_old,   5, 1),
            }

    for key, value in DICT.items():
        ws = wb[key]
        rows = dataframe_to_rows(value[0], index=False, header=False)
        for r_idx, row in enumerate(rows, value[1]):
            for c_idx, val in enumerate(row, value[2]):
                ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(new_name_pred)

    # основной отчёт
    # del sput[sput.columns[-1]]
    # del epivak[epivak.columns[-1]]
    # del covivak[covivak.columns[-1]]
    # del light[light.columns[-1]]
    del revac['SCEP']

    wb = openpyxl.load_workbook(new_name_osn)
    DICT = {
        'Спутник-V':    (sput,    5, 1),
        'ЭпиВакКорона': (epivak,  5, 1),
        'КовиВак':      (covivak, 5, 1),
        'Спутник Лайт': (light,   5, 1),
        'Ревакцинация': (revac,   9, 1),
            }

    for key, value in DICT.items():
        ws = wb[key]
        rows = dataframe_to_rows(value[0], index=False, header=False)
        for r_idx, row in enumerate(rows, value[1]):
            for c_idx, val in enumerate(row, value[2]):
                ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(new_name_osn)

    return new_name_pred + ';' + new_name_osn
