import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql
from pandas import to_datetime


class no_data(Exception):
    pass


def svod_51_covid_19():
    SQL_1 = open('parus/sql/covid_51_svod.sql', 'r').read()
    SQL_2 = open('parus/sql/covid_51_svod_all.sql', 'r').read()

    DF = parus_sql(SQL_1)
    ALL = parus_sql(SQL_2)

    if len(DF) == 0:
        raise no_data('Нет данных за сегодня!')

    try:
        DATE = DF.at[0, 'DAY']
        del DF['DAY']
    except KeyError:
        DATE = to_datetime(ALL['DAY'], format='%d.%m.%Y').max()
        DATE = DATE.strftime('%d.%m.%Y')

    DF = DF.append(DF.sum(numeric_only=True), ignore_index=True)

    DF.loc[len(DF)-1, 'COV_02'] = 'ИТОГО:'

    NEW_NAME = f'/tmp/{str(DATE)}_51_COVID_19_cvod.xlsx'

    shutil.copyfile('help/51_COVID_19_svod.xlsx', NEW_NAME)

    wb = load_workbook(NEW_NAME)

    ws = wb['Свод по МО']
    rows = dataframe_to_rows(DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Свод по всем МО']
    rows = dataframe_to_rows(ALL, index=False, header=False)
    for r_idx, row in enumerate(rows, 4):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME)

    return NEW_NAME
