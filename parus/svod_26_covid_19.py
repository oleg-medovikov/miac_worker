from datetime import datetime, timedelta
import shutil
import openpyxl
import pandas as pd
from glob import glob
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql
from clas import Dir


class my_except(Exception):
    pass


def svod_26_covid_19():
    SQL = open("parus/sql/covid_26_svod.sql", "r").read()

    DF = parus_sql(SQL)
    DATE = DF.at[0, "DAY"]

    del DF["DAY"]

    DF["type"] = "parus"

    # DATE = (datetime.now() - timedelta(days=0)).strftime('%d.%m.%Y')
    # OLD_FILE = Dir.get('punct_zabor') + '/' + DATE + ' Пункты отбора.xlsx'

    FILES = glob(Dir.get("punct_zabor") + "/* Пункты отбора.xlsx")

    try:
        OLD_FILE = FILES[0]
    except IndexError:
        raise my_except("не найден файл старого отчета!")

    values = {6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0}

    try:
        OLD = pd.read_excel(OLD_FILE, skiprows=3, header=None, sheet_name="Соединение")
    except Exception:
        OLD_ = pd.DataFrame()
        OLD = pd.DataFrame()
    else:
        OLD = OLD.loc[~(OLD[2].isnull() & OLD[3].isnull() & OLD[5].isnull())]
        OLD_ = OLD.fillna(value=values).copy()
        del OLD_[0]
        del OLD[0]
        del OLD_[14]
        for _ in [7, 9, 11, 13]:
            OLD_[_] = 0

        OLD_["type"] = "file"

    if len(OLD_.columns) == len(DF.columns):
        OLD_.columns = DF.columns

    NEW_DF = pd.concat([DF, OLD_], ignore_index=True)
    NEW_DF = NEW_DF.drop_duplicates(subset=["LAB_UTR_MO", "ADDR_PZ", "LAB_UTR_02"])

    # DATE = (datetime.now() + timedelta(days=1)).strftime("%d_%m_%Y")

    NEW_NAME = "temp/" + DATE + "_26_COVID_19_cvod.xlsx"

    shutil.copyfile("help/26_COVID_19_svod.xlsx", NEW_NAME)

    wb = openpyxl.load_workbook(NEW_NAME)

    ws = wb["Из паруса"]
    rows = dataframe_to_rows(DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb["Из файла"]
    rows = dataframe_to_rows(OLD, index=False, header=False)
    for r_idx, row in enumerate(rows, 5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb["Соединение"]
    rows = dataframe_to_rows(NEW_DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME)

    return NEW_NAME
