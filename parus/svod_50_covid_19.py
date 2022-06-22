import time, datetime, shutil, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql, covid_sql

def svod_50_covid_19():
    SQL_1 = open('parus/sql/covid_50_polic.sql', 'r').read()
    SQL_2 = open('parus/sql/covid_50_stac.sql', 'r').read()
    SQL_3 = open('parus/sql/mz_50.sql', 'r').read()

    SQL_4 = open('parus/sql/covid_50_polic_old.sql', 'r').read()
    SQL_5 = open('parus/sql/covid_50_stac_old.sql', 'r').read()

    POLIC     = parus_sql(SQL_1)
    STAC      = parus_sql(SQL_2)
    COVID     = covid_sql(SQL_3)
    POLIC_OLD = parus_sql(SQL_4)
    STAC_OLD  = parus_sql(SQL_5)

    DATE = POLIC.at[0,'DAY']

    del POLIC ['DAY']
    del STAC ['DAY']

    COVID_POL = COVID.loc [COVID['Type_Therapy'] == 'Поликлинника']
    
    del COVID_POL ['Type_Therapy']
    del COVID_POL ['Count_70_COVID_week']
    del COVID_POL ['Count_70_COVID']
    del COVID_POL ['Count_70_Pnev_week']
    del COVID_POL ['Count_70_Pnev']

    COVID_STAC = COVID.loc [COVID['Type_Therapy'] == 'Стационар']
    
    del COVID_STAC ['Type_Therapy']

    NEW_NAME = 'temp/50_COVID_19_' + DATE + '_предварительный.xlsx'

    shutil.copyfile('help/50_COVID_19_pred.xlsx', NEW_NAME )

    wb= openpyxl.load_workbook( NEW_NAME ) 

    ws = wb['Разрез МО_ГП']
    rows = dataframe_to_rows(POLIC,index=False, header=False)
    index_col = [2,9,10,11,12,13,14,15,16,17,18,19,20,21]
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)
    
    ws = wb['Разрез МО_стац']
    rows = dataframe_to_rows(STAC,index=False, header=False)
    index_col = [2,13,14,15,16,17,18,19,20,21,22,23,24]
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)
    
    ws = wb['Пред.отч_разрез МО_ГП']
    rows = dataframe_to_rows(POLIC_OLD,index=False, header=False)
    index_col = [2,9,10,11,12,13,14,15,16,17,18,19,20,21]
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)
    
    ws = wb['Пред.отч_разрез МО_стац']
    rows = dataframe_to_rows(STAC_OLD,index=False, header=False)
    index_col = [2,13,14,15,16,17,18,19,20,21,22,23,24]
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)
    
    ws = wb['ФР_ГП']
    rows = dataframe_to_rows(COVID_POL,index=False, header=False)
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['ФР_стац']
    rows = dataframe_to_rows(COVID_STAC,index=False, header=False)
    for r_idx, row in enumerate(rows,7):  
        for c_idx, value in enumerate(row,2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( NEW_NAME ) 

    return NEW_NAME
