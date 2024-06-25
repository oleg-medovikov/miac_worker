from datetime import datetime, timedelta
from pandas import concat, to_numeric
import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql

from .svod_40_sql import SQL_VACHIN, SQL_REVAC_MO

LIST_POK = [
    "DAY",
    "TYPE",
    "distr",
    "org",
    "_01",
    "_02",
    "_03",
    "_04",
    "_05",
    "_06",
    "_07",
    "_08",
    "_09",
    "_10",
    "_11",
    "_12",
    "_13",
    "_14",
    "_15",
    "_16",
    "_17",
    "_18",
    "_19",
    "_20",
    "_21",
    "_22",
    "_23",
    "_24",
    "_25",
    "_26",
    "_27",
    "_28",
    "_29",
    "_30",
    "_31",
    "_32",
]

POKAZATELS = ["KOVIVAC", "KONVASEL", "CPUTNIKL", "CPUTNIKV", "EPIVAK"]


def svod_40_covid_new():
    "Новый свод 40 ковида"
    DICT = {}

    # Вытаскиваем вакцины всем скопом и разбиваем на таблички
    DF = parus_sql(SQL_VACHIN)
    DF = DF.pivot_table(
        index=["DAY", "ORGANIZATION", "TYPE", "ROW_INDEX"],
        columns=["POKAZATEL"],
        values="VALUE",
        aggfunc="first",
    )
    DF = DF.reset_index()

    for col in DF.columns:
        try:
            DF[col] = to_numeric(DF[col])
        except ValueError:
            continue

    new_name_pred = f'/tmp/40_COVID_19_БОТКИНА_{DF["DAY"].min()}__{DF["DAY"].max()}_предварительный.xlsx'

    new_name_osn = (
        f'/tmp/40_COVID_19_БОТКИНА_{DF["DAY"].min()}__{DF["DAY"].max()}_основной.xlsx'
    )

    for POKAZATEL in POKAZATELS:
        # Формируем список колонок
        COLUMNS = [x.replace("_", POKAZATEL + "_") for x in LIST_POK]
        # Добавляем недостающие
        for COL in COLUMNS:
            if COL not in DF.columns:
                DF[COL] = ""

        PART = DF[COLUMNS].copy()
        # проставляем назхвания организаций

        PART.loc[(PART["TYPE"] == "Медицинская организация"), POKAZATEL + "_02"] = PART[
            "org"
        ]
        PART.loc[(PART["TYPE"] == "Медицинская организация"), POKAZATEL + "_01"] = PART[
            "distr"
        ]

        del PART["org"]
        del PART["distr"]

        PART.loc[
            (PART[POKAZATEL + "_01"].isnull()) & (PART["TYPE"] == "Пункт вакцинации"),
            POKAZATEL + "_01",
        ] = (
            DF[POKAZATEL + "_02"].str.split().str[0]
        )

        # какие-то махинации с вырезкой района из названия ТВСП
        PART.loc[
            ~(PART[POKAZATEL + "_01"].isnull())
            & ~(PART[POKAZATEL + "_01"].str.contains("район", na=False)),
            POKAZATEL + "_02",
        ] = (
            PART[POKAZATEL + "_02"].str.split(n=1).str[1]
        )
        PART = PART.loc[~PART[POKAZATEL + "_02"].isnull()]

        DISTRICTS = [
            str(x).replace(" район", "") for x in PART[POKAZATEL + "_01"].unique()
        ]
        for DIST in DISTRICTS:
            PART[POKAZATEL + "_02"] = PART[POKAZATEL + "_02"].str.replace(str(DIST), "")
        PART[POKAZATEL + "_02"] = PART[POKAZATEL + "_02"].str.lstrip()

        # особенность спутника лайта - итого по первой и второй вакцине равны
        if POKAZATEL == "CPUTNIKL":
            PART["CPUTNIKL_26"] = PART["CPUTNIKL_24"]
            PART["CPUTNIKL_27"] = PART["CPUTNIKL_25"]
        # конвасел оказывается тоже однокомпонетный!
        if POKAZATEL == "KONVASEL":
            PART["KONVASEL_26"] = PART["KONVASEL_24"]
            PART["KONVASEL_27"] = PART["KONVASEL_25"]

        # запись итоговых данных
        DICT[POKAZATEL] = PART.loc[PART["DAY"] == PART["DAY"].max()]
        if PART["DAY"].min() != PART["DAY"].max():
            DICT[POKAZATEL + "_OLD"] = PART.loc[PART["DAY"] == PART["DAY"].min()]

    # Вытаскиваем ревакцинацию за МО
    list_ = []
    D_VAC = {
        "1": "Всего",
        "2": "Гам-КОВИД-Вак (Спутник-V)",
        "3": "КовиВак",
        "4": "ЭпиВакКорона",
        "5": "Спутник Лайт",
        "6": "Конвасэл",
    }
    for i in range(1, 7):
        SQL_ = SQL_REVAC_MO.replace("индекс", str(i))
        DF = parus_sql(SQL_)
        list_.append(DF)

    DF = concat(list_, ignore_index=True)
    DF["TYPEVACINE"] = DF["INDX"].map(D_VAC)
    DF["SCEP"] = DF["ORGANIZATION"] + " " + DF["TYPEVACINE"]

    # TVSP = parus_sql(SQL_REVAC_TVSP)
    # DF = concat([DF, TVSP], ignore_index=True)
    # особенность спутника лайта
    DF.loc[DF["INDX"] == "5", "POK_18"] = DF["POK_16"]
    DF.loc[DF["INDX"] == "5", "POK_19"] = DF["POK_17"]
    DF.loc[DF["INDX"] == "5", "POK_16"] = 0
    DF.loc[DF["INDX"] == "5", "POK_17"] = 0
    # особенность конвасела
    DF.loc[DF["INDX"] == "6", "POK_18"] = DF["POK_16"]
    DF.loc[DF["INDX"] == "6", "POK_19"] = DF["POK_17"]
    DF.loc[DF["INDX"] == "6", "POK_16"] = 0
    DF.loc[DF["INDX"] == "6", "POK_17"] = 0

    DF = DF.sort_values(["ORGANIZATION", "INDX"])

    DICT["REVAC"] = DF.loc[DF["DAY"] == DF["DAY"].max()].copy()
    if DF["DAY"].min() != DF["DAY"].max():
        DICT["REVAC_OLD"] = DF.loc[DF["DAY"] == DF["DAY"].min()].copy()

    shutil.copyfile("/mnt/COVID-списки/bot/40_COVID_19_pred_new.xlsx", new_name_pred)
    shutil.copyfile("/mnt/COVID-списки/bot/40_COVID_19_osn_new.xlsx", new_name_osn)

    # Записываем данные в предварительный файл
    wb = openpyxl.load_workbook(new_name_pred)
    D_PRINT = {
        "Спутник-V": ("CPUTNIKV", 5, 1),
        "Вчера_Спутник": ("CPUTNIKV_OLD", 5, 1),
        "ЭпиВакКорона": ("EPIVAK", 5, 1),
        "КовиВак": ("KOVIVAC", 5, 1),
        "Конвасэл": ("KONVASEL", 5, 1),
        "Спутник Лайт": ("CPUTNIKL", 5, 1),
        "Ревакцинация": ("REVAC", 10, 1),
        "Вчера_КовиВак": ("KOVIVAC_OLD", 5, 1),
        "Вчера_ЭпиВак": ("EPIVAK_OLD", 5, 1),
        "Вчера_Спутник Лайт": ("CPUTNIKL_OLD", 5, 1),
        "Вчера_ревакцин": ("REVAC_OLD", 10, 1),
        "Вчера_Конвасэл": ("KONVASEL_OLD", 5, 1),
    }

    for key, value in D_PRINT.items():
        ws = wb[key]
        try:
            DATA = DICT[value[0]]
        except KeyError:
            continue

        del DATA["DAY"]

        for i in range(value[1], ws.max_row):
            if ws.cell(row=i, column=1).value is None:
                rowNum = i
                break

        rows = dataframe_to_rows(DATA, index=False, header=False)
        for r_idx, row in enumerate(rows, rowNum):
            for c_idx, val in enumerate(row, value[2]):
                ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(new_name_pred)

    wb = openpyxl.load_workbook(new_name_osn)
    D_PRINT = {
        "Спутник-V": ("CPUTNIKV", 5, 1),
        "ЭпиВакКорона": ("EPIVAK", 5, 1),
        "КовиВак": ("KOVIVAC", 5, 1),
        "Конвасэл": ("KONVASEL", 5, 1),
        "Спутник Лайт": ("CPUTNIKL", 5, 1),
        "Ревакцинация": ("REVAC", 10, 1),
    }

    for key, value in D_PRINT.items():
        ws = wb[key]
        DATA = DICT[value[0]]

        for i in range(value[1], ws.max_row):
            if ws.cell(row=i, column=1).value is None:
                rowNum = i
                break

        rows = dataframe_to_rows(DATA, index=False, header=False)
        for r_idx, row in enumerate(rows, rowNum):
            for c_idx, val in enumerate(row, value[2]):
                ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(new_name_osn)

    return new_name_pred + ";" + new_name_osn
