import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql
import pandas as pd

from system import dict_ogrn_frmoname


def distant_consult():
    SQL = open('parus/sql/distant_svod.sql', 'r').read()

    DF = parus_sql(SQL)
    DICT_ORG = dict_ogrn_frmoname()

    for i in DF.index:
        DF.loc[i, 'POK02'] = DICT_ORG.get(
            int(DF.at[i, 'ORGANIZATION']),
            DF.at[i, 'POK02']
            )

    DATE = DF.at[0, 'DAY']
    del DF['DAY']

    NEW_NAME = 'temp/Дистанц_Консультации_' + DATE + '.xlsx'

    shutil.copyfile('help/distant_consult.xlsx', NEW_NAME)

    OLD = pd.read_excel('help/distant_consult.xlsx', sheet_name='эталон')

    DOLG = pd.DataFrame(columns=["ORGANIZATION"])

    for i in range(len(OLD)):
        ORG = OLD.at[i, 'Полное наименование МО (из ФРМО)']
        if ORG not in DF["POK02"].unique():
            col = 'Краткое наименование МО (для вывода должников)'
            DOLG.loc[len(DOLG), "ORGANIZATION"] = OLD.at[i, col]

    del DF['ORGANIZATION']

    wb = openpyxl.load_workbook(NEW_NAME)

    ws = wb['svod']

    rows = dataframe_to_rows(DF, index=False, header=False)
    index_col = [
        1, 2, 20, 21, 22, 23, 24, 25, 26, 27, 28,
        29, 30, 31, 32, 33, 34, 35, 36, 37, 38
        ]
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 0):
            ws.cell(row=r_idx, column=index_col[c_idx], value=value)

    ws = wb['dolg']

    rows = dataframe_to_rows(DOLG, index=False, header=False)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME)

    CSV = pd.read_csv('help/dist_cons.csv', sep=';')

    CSV = CSV.drop(0)

    k = 0
    for i in index_col:
        try:
            for index in range(len(DF)):
                CSV.loc[index, CSV.columns[i-1]] = str(DF.at[
                                        index,
                                        DF.columns[k]
                                    ]).replace('.0', '')
        except:
            break
        else:
            k += 1

    CSV_FILE = 'temp/шаблон_для_загрузки.csv'

    CSV.to_csv(CSV_FILE,  sep=';', index=False, encoding='cp1251')

    return NEW_NAME + ';' + CSV_FILE
