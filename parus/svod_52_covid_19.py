import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

from base import parus_sql

DICT_ORGS = {
    'ООО "Американская медицинская клиника"': '22.03.2023',
    'ООО «Веда»': '22.03.2023',
    'ООО «Дельфин»': '22.03.2023',
    'ООО «Делар»': '22.03.2023',
    'ООО «Ладья»': '22.03.2023',
    'ООО "НК"АБИА"': '22.03.2023',
    'ООО «ОСМА КЛИНИК»': '22.03.2023',
    'ООО "Лахта Клиника"': '22.03.2023',
    'АНО "МИЦЛТ "Юнинова""':  '22.03.2023',
    'ООО «Клиника Доктора Пеля»':  '22.03.2023',
    'ООО "М89"':  '22.03.2023',
    'ООО "Морской Медицинский Центр"':  '22.03.2023',
    'ООО "МЦ "XXI век" СЗ"':  '22.03.2023',
    'ООО "НК"АБИА"':  '22.03.2023',
    'ООО "СОГАЗ Профмедицина"':  '22.03.2023',
    'СПб ГБУЗ "Городская больница №40"':  '22.03.2023',
    'СПб ГБУЗ "Городская клиническая больница №31"':  '22.03.2023',
    'СПб ГБУЗ "Городская поликлиника №22"':  '22.03.2023',
    'СПб ГБУЗ "Городская поликлиника №27"':  '22.03.2023',
    'СПб ГБУЗ "Городская поликлиника №3"':  '22.03.2023',
    'СПб ГБУЗ "Городская поликлиника №4"':  '22.03.2023',
    'СПб ГБУЗ "Городская поликлиника №43"':  '22.03.2023',
    'СПб ГБУЗ "Клиническая инфекционная больница им. С.П. Боткина"':
        '22.03.2023',
    'СПб ГМУ им.И.П.Павлова "Городская поликлиника №31"':  '22.03.2023',
    'ФГБОУ ВО СЗГМУ им. И.И. МЕЧНИКОВА МИНЗДРАВА РОССИИ':  '22.03.2023',
    'ФГБОУ ВО СПБГПМУ МИНЗДРАВА РОССИИ':  '22.03.2023',
    'АО «Адмиралтейские верфи»': '06.04.2023',
    'ООО «Единый Медицинский Центр»': '06.04.2023',
    'ООО «МедМигСервис»': '08.04.2023',
    'ООО "МедЭксперт"': '08.04.2023',
    'ООО "ИНВИТРО СПб"': '08.04.2023',
    'ООО «МЕДЕФ»': '08.04.2023',
    'ООО «МЕДИКУМ»': '08.04.2023',
    'ООО "Медилюкс-ТМ"': '08.04.2023',
    'ООО "Медицинские Услуги"': '08.04.2023',
    'ООО «Медосмотр»': '08.04.2023',
    'ООО "Медси Санкт-Петербург"': '08.04.2023',
    'ООО "Международный медицинский центр "СОГАЗ"': '08.04.2023',
    'ООО “МЦ “Мария”': '08.04.2023',
    'ООО «Немецкая семейная  клиника»': '08.04.2023',
    'ООО "РеаСанМед"': '08.04.2023',
    'ООО "ЦЗЗ "Благомед"': '08.04.2023',
        }


def svod_52_covid_19():
    ORGANIZATIONS = ''
    for key, value in DICT_ORGS.items():
        ORGANIZATIONS += f"\n\t\tOR (a.AGNNAME = '{key}' \
            AND r.BDATE = TO_DATE('{value}','DD.MM.YYYY'))"

    SQL_1 = open('parus/sql/covid_52_svod.sql', 'r').read()
    SQL_2 = open('parus/sql/covid_52_svod_old.sql', 'r').read()

    SQL_1 = SQL_1.replace('ORGANIZATIONS', ORGANIZATIONS)
    SQL_2 = SQL_2.replace('ORGANIZATIONS', ORGANIZATIONS)

    DF = parus_sql(SQL_1)
    OLD = parus_sql(SQL_2)

    DATE = DF.at[0, 'DAY']

    del DF['DAY']
    del DF['ORGANIZATION']
    del OLD['DAY']
    del OLD['ORGANIZATION']

    NEW_NAME_1 = 'temp/' + DATE + '_52_COVID_19_pred.xlsx'
    NEW_NAME_2 = 'temp/' + DATE + '_52_COVID_19_osn.xlsx'

    shutil.copyfile('help/52_COVID_19_pred.xlsx', NEW_NAME_1)
    shutil.copyfile('help/52_COVID_19_osn.xlsx',  NEW_NAME_2)

    wb = openpyxl.load_workbook(NEW_NAME_1)

    ws = wb['52 COVID']
    rows = dataframe_to_rows(DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 11):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Вчера']
    rows = dataframe_to_rows(OLD, index=False, header=False)
    for r_idx, row in enumerate(rows, 11):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME_1)

    wb = openpyxl.load_workbook(NEW_NAME_2)

    ws = wb['52 COVID']
    rows = dataframe_to_rows(DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 11):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME_2)

    return NEW_NAME_1 + ';' + NEW_NAME_2
