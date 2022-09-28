from openpyxl.utils.dataframe import dataframe_to_rows
import shutil, openpyxl
import pandas as pd
from base import parus_sql
from clas import Dir

ROWS = {
    '_01_' : '1) Т90 (Последствия травм головы)',
    '_02_' : '2) Т90.4 (Последствия травмы глаза, окологлазничной области)',
    '_03_' : '3) Т91 (Последствия травм шеи и туловища)',
    '_04_' : '4) Т92 (Последствия травм верхней конечности)',
    '_05_' : '5) Т92.6 (Последствия размозжения и травматической ампутации верхней конечности)',
    '_06_' : '6) Т93 (Последствия травм нижней конечности)',
    '_07_' : '7) Т93.6 (Последствия размозжения и травматической ампутации нижней конечности)',
    '_08_' : '8) Т94 (Последствия травм, захватывающих несколько областей тела, и травм неуточненной локализации)',
    '_09_' : '9) Т05 (Травматические ампутации, захватывающие несколько областей тела)'
}

COLS = {
        '_01' : ' 1',
        '_02' : ' 2',
        '_03' : ' 3',
        '_05' : ' 4',
        '_06' : ' 5',
        '_08' : ' 6',
        '_09' : ' 7',
        '_11' : ' 8',
        '_12' : ' 9',
        '_14' : '10',
        '_15' : '11',
        '_17' : '12',
        '_18' : '13',
        '_20' : '14',
        '_21' : '15',
        '_23' : '16',
        '_24' : '17',
        '_26' : '18',
        '_27' : '19',
        '_29' : '20',
        '_30' : '21',
}

def mp_travm_v_hode_co():
    sql = open('parus/sql/mp_travm_v_hode_co.sql', 'r').read()

    DF = parus_sql(sql)

    DATE = DF['DAY'].unique()[0]

    for key, value in ROWS.items():
        DF.loc[DF['POKAZATEL'].str.contains(key), 'ROWS'] = value

    
    for key, value in COLS.items():
        DF.loc[DF['POKAZATEL'].str.endswith(key), 'COLS'] = value


    O = DF.pivot_table(index=['ORGANIZATION','ROWS'], columns=['COLS'],values=['VALUE'], aggfunc='first').stack(0)

    
    O.reset_index(inplace=True)  

    del O [' 1']
    del O ['level_2']

    for org in O['ORGANIZATION'].unique():
        district = O.loc[(~(O[' 2'].isnull()) & (O['ORGANIZATION'] == org)), ' 2'].iat[0]
        O.loc[((O[' 2'].isnull()) & (O['ORGANIZATION'] == org)), ' 2'] = district

    names = ['ORGANIZATION', ' 2', 'ROWS'] + list(O.columns[3:])
    O = O.reindex(columns=names)

    S = O.copy()
    for col in S.columns[3:]:
        S[col] = pd.to_numeric( S[col] )

    S = S.groupby([' 2', 'ROWS' ], as_index=False).sum()


    NEW_NAME = 'temp/' + DATE + '_МП_травм_в_ходе_СО.xlsx'

    shutil.copyfile('help/mp_travm_v_hode_co.xlsx', NEW_NAME )

    wb= openpyxl.load_workbook( NEW_NAME )

    ws = wb['свод']
    rows = dataframe_to_rows(O,index=False, header=False)
    for r_idx, row in enumerate(rows,12):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
 
    ws = wb['районы']
    rows = dataframe_to_rows(S,index=False, header=False)
    for r_idx, row in enumerate(rows,12):  
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
 
    
    wb.save( NEW_NAME )

    return NEW_NAME


