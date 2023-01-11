import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from base import parus_sql


def name_test_system(STR: str) -> str:
    "МО не могут написать название тест систем, чистим"
    word = STR.replace('"', '')
    word = word.lower().replace(' ', '')
    return {
        'sars' in word and 'cov' in word: 'SARS-CoV-2 Ag-ИМБИАН-ИХА',
        'sars' in word and 'ag' in word: 'SARS-CoV-2 Ag-ИМБИАН-ИХА',
        'sara-cov' in word: 'SARS-CoV-2 Ag-ИМБИАН-ИХА',
        'имбиан' in word: 'SARS-CoV-2 Ag-ИМБИАН-ИХА',
        'imbian' in word:  'SARS-CoV-2 Ag-ИМБИАН-ИХА',
        'sar-cov-2' in word: 'SARS-CoV-2 Ag-ИМБИАН-ИХА',
        'rapid' in word: 'COVID 19 Antigen Rapid Test kit',
        'covid' in word and 'antigen' in word:
            'COVID 19 Antigen Rapid Test kit',
        'covid' in word and 'антиген' in word:
            'COVID 19 Antigen Rapid Test kit',
        'иха' in word: 'SARS-CoV-2 Ag-ИМБИАН-ИХА',
    }.get(True,  STR.replace('"', ''))


def svod_54_covid_new2():
    SQL = open('parus/sql/covid_54_new2.sql', 'r').read()

    DF = parus_sql(SQL)

    DATE = datetime.today().strftime('%d.%m.%Y')  # DF.at[0, 'DAY']

    del DF['DAY']

    DF['POK_3'] = DF['POK_3'].apply(name_test_system)
    DF['POK_4'] = DF['POK_4'].str.replace('.', ',')

    DF = DF.groupby(by=[
        'ORGANIZATION',
        'OGRN',
        'POK_3',
        'POK_4'
        ], as_index=False).sum()

    NEW_NAME = f'temp/{DATE}_54_COVID_19_new2.xlsx'

    shutil.copyfile('help/54_COVID_19_new2.xlsx', NEW_NAME)

    wb = load_workbook(NEW_NAME)

    ws = wb['свод']
    rows = dataframe_to_rows(DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME)

    return NEW_NAME
