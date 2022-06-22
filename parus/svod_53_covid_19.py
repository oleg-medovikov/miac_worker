import time, datetime, shutil, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql

def svod_53_covid_19():
    SQL_1 = open('parus/sql/covid_53_svod.sql',     'r').read()
    SQL_2 = open('parus/sql/covid_53_svod_old.sql', 'r').read()
    
    DF  = parus_sql( SQL_1 )
    OLD = parus_sql( SQL_2 )
    
    DATE = DF.at[0,'DAY']
    del DF['DAY']
    
    SM = DF.groupby(by="ORGANIZATION",as_index=False).sum()
    
    SM ['POK03'] = SM ['ORGANIZATION']

    SM ['TYPE'] = 'Медицинская организация'


    for i in range(len(SM)):
        k = len(DF)
        for col in DF.columns:
            try:
                DF.loc[k,col] = SM.at[i,col]
            except:
                pass

    DF = DF.sort_values(by=["ORGANIZATION", "POK02"],
            na_position='first',
            ignore_index=True).fillna('')
    
    del DF ['ORGANIZATION']

    # вчерашнее 
    del OLD['DAY']

    SM = OLD.groupby(by="ORGANIZATION",as_index=False).sum()
    
    SM ['POK03'] = SM ['ORGANIZATION']
    SM ['TYPE'] = 'Медицинская организация'
    
    for i in range(len(SM)):
        k = len(OLD)
        for col in OLD.columns:
            try:
                OLD.loc[k,col] = SM.at[i,col]
            except:
                pass

    OLD = OLD.sort_values(by=["ORGANIZATION", "POK02"],
            na_position='first',
            ignore_index=True).fillna('')
    
    del OLD ['ORGANIZATION']
    # ====

    
    NEW_NAME_1 = 'temp/53_COVID_БОТКИНА_' + DATE + '.xlsx'
    NEW_NAME_2 = 'temp/53_COVID_БОТКИНА_' + DATE + '_основной.xlsx'

    shutil.copyfile('help/53_COVID_19_svod.xlsx', NEW_NAME_1)
    shutil.copyfile('help/53_COVID_19.xlsx',      NEW_NAME_2)

    wb= openpyxl.load_workbook( NEW_NAME_1 )

    ws = wb['Спутник-М']
    rows = dataframe_to_rows(DF,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Вчера_Спутник-М']
    rows = dataframe_to_rows(OLD,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( NEW_NAME_1 )

    wb= openpyxl.load_workbook( NEW_NAME_2 )

    ws = wb['Спутник-М']
    rows = dataframe_to_rows(DF,index=False, header=False)
    for r_idx, row in enumerate(rows,5):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( NEW_NAME_2 )

    return NEW_NAME_1 + ';' + NEW_NAME_2
