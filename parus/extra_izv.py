import shutil
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from base import parus_sql, miac_ds_sql
from conf import REGIZ_AUTH
import requests


def extra_izv():
    SQL_1 = open("parus/sql/extra_izv.sql", "r").read()
    SQL_3 = open("parus/sql/extra_izv_regiz.sql", "r").read()

    # SQL_2 = open('parus/sql/extra_izv_dolg.sql', 'r').read()

    DF = parus_sql(SQL_1)
    REG = miac_ds_sql(SQL_3)
    # OLD  = parus_sql( SQL_2 )

    URL = f"https://regiz.gorzdrav.spb.ru/N3.BI/getDData?id=1440&auth={REGIZ_AUTH}"
    req = requests.get(URL)
    if req.status_code == 200:
        ANAL = pd.DataFrame(data=req.json())
        ANAL = ANAL[
            ["month", "oid", "district", "MO", "Doc_Count", "Remd_Count", "Sign_Count"]
        ]
    else:
        ANAL = pd.DataFrame()

    DATE = DF.at[0, "DAY"]

    del DF["DAY"]

    OLD = pd.read_excel(
        "help/extra_izv.xlsx", sheet_name="list", usecols="B", names=["ORGANIZATION"]
    )

    DOLG = pd.DataFrame(columns=["ORGANIZATION"])

    for ORG in OLD["ORGANIZATION"].unique():
        if ORG not in DF["POK01"].unique():
            DOLG.loc[len(DOLG), "ORGANIZATION"] = ORG

    NEW_NAME = "temp/Экстренные_извещения_" + DATE + ".xlsx"

    shutil.copyfile("help/extra_izv.xlsx", NEW_NAME)

    wb = openpyxl.load_workbook(NEW_NAME)
    ws = wb["svod"]
    rows = dataframe_to_rows(DF, index=False, header=False)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    rows = dataframe_to_rows(DOLG, index=False, header=False)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, value in enumerate(row, 6):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb["regiz"]
    rows = dataframe_to_rows(REG, index=False, header=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb["anal"]
    rows = dataframe_to_rows(ANAL, index=False, header=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(NEW_NAME)

    return NEW_NAME
