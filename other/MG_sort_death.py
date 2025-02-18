import datetime
import glob
import pandas as pd
import openpyxl
from collections import namedtuple
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
from base import covid_sql


def MG_sort_death():
    def search_mo(street, house):
        """Search for Medical Organization (MO) based on street and house."""
        for mo in mo_org:
            if mo.Street == street and mo.House == house:
                return mo.Name_MO
        return "Не найдено МО"

    # Get today's date in the required format
    date_otch = datetime.datetime.today().strftime("%d.%m.%Y")

    # Locate the Excel file for today's report
    file_path = f"/mnt/COVID-списки/ДарьинаМГ/УМСРС сортировка умерших/[!~]*Умершие от Covid-19*{date_otch}*.xlsx"
    excel_files = glob.glob(file_path)

    if not excel_files:
        return "Я не нашёл файлик за сегодня!"

    # Define columns to use from the Excel file
    cols = [
        "№ п/п",
        "Возраст",
        "Субъект",
        "Улица смерти",
        "Дом смерти",
        "Краткое наименование",
        "Место смерти",
    ]

    # Load the data from the Excel file
    df = pd.read_excel(excel_files[0], header=1, usecols=cols)
    df = df[df["№ п/п"].notnull() & (df["№ п/п"] != 0)]
    df.reset_index(drop=True, inplace=True)

    # Fetch MO data from the database
    mo = namedtuple("mo", ["Name_MO", "Street", "House"])
    sql = """
    SELECT [Name_MO], [Street], [House]
    FROM [COVID].[Nsi].[Address_MO]
    """

    adress = covid_sql(sql)  # это DataFrame
    mo_org = [
        mo(row["Name_MO"], row["Street"], row["House"]) for _, row in adress.iterrows()
    ]

    # Assign Name_MO based on conditions
    df.loc[~df["Место смерти"].isin(["в стационаре"]), "Name_MO"] = "БСМЭ\ПАБ"
    for idx, row in df[df["Место смерти"].isin(["в стационаре"])].iterrows():
        mo_name = search_mo(row["Улица смерти"], row["Дом смерти"])
        df.at[idx, "Name_MO"] = (
            mo_name if mo_name != "Не найдено МО" else row["Краткое наименование"]
        )

    # Generate summary report
    otchet = (
        df[df["Субъект"] != "Ленинградская обл"]
        .groupby(by="Name_MO", as_index=False)
        .count()[["Name_MO", "№ п/п"]]
        .rename(columns={"Name_MO": "Медицинская организация", "№ п/п": "Всего СПб"})
    )

    leningrad_data = (
        df[df["Субъект"] == "Ленинградская обл"]
        .groupby(by="Name_MO", as_index=False)
        .count()[["Name_MO", "№ п/п"]]
        .rename(columns={"Name_MO": "Медицинская организация", "№ п/п": "Всего ЛО"})
    )

    otchet = otchet.merge(leningrad_data, how="outer", on="Медицинская организация")

    # Add age-based statistics
    status_labels = [
        "Дети до 18 СПб",
        "Взрослые до 65 СПб",
        "Пенсионеры после 65 СПб",
        "Дети до 18 ЛО",
        "Взрослые до 65 ЛО",
        "Пенсионеры после 65 ЛО",
    ]
    age_thresholds = [18, 65, 150]

    for label in status_labels:
        otchet[label] = 0

    for _, row in df.iterrows():
        subject = row["Субъект"]
        age = int(row["Возраст"])
        mo_name = row["Name_MO"]

        if subject != "Ленинградская обл":
            if age < age_thresholds[0]:
                otchet.loc[
                    otchet["Медицинская организация"] == mo_name, status_labels[0]
                ] += 1
            elif age < age_thresholds[1]:
                otchet.loc[
                    otchet["Медицинская организация"] == mo_name, status_labels[1]
                ] += 1
            elif age < age_thresholds[2]:
                otchet.loc[
                    otchet["Медицинская организация"] == mo_name, status_labels[2]
                ] += 1
        else:
            if age < age_thresholds[0]:
                otchet.loc[
                    otchet["Медицинская организация"] == mo_name, status_labels[3]
                ] += 1
            elif age < age_thresholds[1]:
                otchet.loc[
                    otchet["Медицинская организация"] == mo_name, status_labels[4]
                ] += 1
            elif age < age_thresholds[2]:
                otchet.loc[
                    otchet["Медицинская организация"] == mo_name, status_labels[5]
                ] += 1

    # Save the report to an Excel file
    output_file = f"/mnt/COVID-списки/ДарьинаМГ/УМСРС сортировка умерших/Свод по возрастам {date_otch}.xlsx"
    shutil.copyfile("help/ШаблонМГ.xlsx", output_file)

    wb = openpyxl.load_workbook(output_file)
    ws = wb["Свод по возрастам"]
    rows = dataframe_to_rows(otchet, index=False, header=False)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_file)

    # Create a transposed summary
    new = pd.DataFrame()
    for _, row in otchet.iterrows():
        k = len(new)
        new.loc[k, "Медицинская организация"] = row["Медицинская организация"]
        new.loc[k, "Всего СПб"] = row["Всего СПб"]
        new.loc[k, "Всего ЛО"] = row["Всего ЛО"]

        for i, label in enumerate(status_labels[:3], start=1):
            new.loc[k + i, "Медицинская организация"] = label.split()[0]
            new.loc[k + i, "Всего СПб"] = row[label] if "СПб" in label else None
            new.loc[k + i, "Всего ЛО"] = row[label] if "ЛО" in label else None

    # Save the transposed summary to the same Excel file
    ws = wb["Перевернутый свод"]
    rows = dataframe_to_rows(new, index=False, header=False)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_file)

    return f"Сгенерирован файл {output_file.split('/')[-1]}"
