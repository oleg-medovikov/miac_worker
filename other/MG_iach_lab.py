from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from clas import Dir
import pandas as pd
from glob import glob
import shutil


def MG_iach_lab():
    """мы тут переводим организации в районы по ведомственной принадлежности"""
    path = Dir.get("MG_iach_lab")
    try:
        file_path = glob(path + "за_*.xlsx")[0]
    except:
        raise ValueError("Не найден файл!")
    else:
        new_file = "/tmp/" + file_path.rsplit("/", 1)[1]
        shutil.copyfile(file_path, new_file)

    df = pd.read_excel(new_file, sheet_name="Свод", header=3, dtype=str)
    m = pd.read_excel(path + "members.xlsx", dtype=str)
    member_ = m.set_index("id")["member"].to_dict()

    df.columns = range(len(df.columns))
    df = df.loc[~df[1].isnull()]
    df[2] = df[1].map(member_)
    df = df.fillna(0)
    del df[0]
    del df[1]
    for col in df.columns:
        if col == 2:
            continue
        df[col] = pd.to_numeric(df[col], errors="coerce")

    member_sort = {
        1: "Василеостровский",
        2: "Колпинский",
        3: "Красногвардейский",
        4: "Красносносельский",
        5: "Курортный",
        6: "Кировский",
        7: "Московский",
        8: "Невский",
        9: "Петроградский",
        10: "Петродворцовый",
        11: "Приморский",
        12: "КЗ",
        13: "РПН",
        14: "Федералы",
        15: "Частные",
    }

    o = df.groupby(2, as_index=False).sum().copy()

    o[1] = o[2].map({v: k for k, v in member_sort.items()})
    o = o.sort_values(by=1)
    o = o[range(1, len(o.columns))]
    d = o.copy()

    itog_row = pd.DataFrame([d.sum(numeric_only=True)], index=[0])

    df = pd.concat([itog_row, o]).reset_index(drop=True)

    df.loc[0, 1] = ""
    df.loc[0, 2] = "Итого"

    wb = load_workbook(new_file)

    if "Районы" in wb.sheetnames:
        del wb["Районы"]

    ws_rayony = wb.copy_worksheet(wb["Свод"])
    ws_rayony.title = "Районы"

    # Удаляем строки ниже 4-й
    if ws_rayony.max_row > 4:
        ws_rayony.delete_rows(5, ws_rayony.max_row - 4)

    # Вставляем данные из DataFrame после 4 строк
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), start=5
    ):
        for c_idx, value in enumerate(row, start=2):
            ws_rayony.cell(row=r_idx, column=c_idx, value=value)

    wb.save(new_file)
    return new_file
