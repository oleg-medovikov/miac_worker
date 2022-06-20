import time, datetime, shutil, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from base import parus_sql

def extra_izv():
    SQL_1 = open('func/parus/sql/extra_izv.sql', 'r').read()
    SQL_2 = open('func/parus/sql/extra_izv_dolg.sql', 'r').read()

    DF   = parus_sql( SQL_1 )
    OLD  = parus_sql( SQL_2 )

    DATE = DF.at[ 0, 'DAY' ]

    del DF ['DAY']

    DOLG = pd.DataFrame(columns=["ORGANIZATION"])

    for ORG in OLD['ORGANIZATION'].unique():
        if not ORG in DF['POK01'].unique():
            DOLG.loc[len(DOLG), 'ORGANIZATION' ] = ORG

    NEW_NAME = 'Экстренные_извещения_' + DATE + '.xlsx'

    shutil.copyfile( 'help/extra_izv.xlsx', NEW_NAME )

    wb = openpyxl.load_workbook ( NEW_NAME )
    
    ws = wb['svod']
    rows = dataframe_to_rows(DF,index=False, header=False)
    for r_idx, row in enumerate(rows,3):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)

    rows = dataframe_to_rows(DOLG,index=False, header=False)
    for r_idx, row in enumerate(rows,3):
        for c_idx, value in enumerate(row, 6):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( NEW_NAME )

    return NEW_NAME


