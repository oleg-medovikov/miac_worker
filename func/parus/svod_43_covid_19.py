import time, datetime, shutil, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql

def svod_43_covid_19():
    SQL_1 = open('func/parus/sql/covid_43_svod.sql', 'r').read()
    
    if datetime.datetime.today().weekday() == 0:
        SQL_2 = open('func/parus/sql/covid_43_svod_old3.sql', 'r').read()
    else:
        SQL_2 = open('func/parus/sql/covid_43_svod_old1.sql', 'r').read()


    DF      = parus_sql(SQL_1)
    DF_OLD  = parus_sql(SQL_2)

    DF.index     = range(1,len(DF) + 1 )
    DF_OLD.index = range(1,len(DF_OLD) + 1 )

    DATE = datetime.datetime.now().strftime('%d_%m_%Y')

    NEW_NAME = 'temp/' + DATE + '_43_COVID_19_cvod.xlsx'

    shutil.copyfile('help/43 COVID 19.xlsx', NEW_NAME)

    wb= openpyxl.load_workbook( NEW_NAME )
    
    ws = wb['Разрез по МО']
    rows = dataframe_to_rows(DF,index=True, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save( shablon_path  + '/' + new_name) 
    
    ws = wb['Вчера']
    rows = dataframe_to_rows(DF_OLD,index=True, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( NEW_NAME )
    
    return NEW_NAME



