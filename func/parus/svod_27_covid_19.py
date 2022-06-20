import time, datetime, shutil, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql

def svod_27_covid_19():
    SQL_1 = open('func/parus/sql/covid_27_svod.sql'    , 'r').read()
    SQL_2 = open('func/parus/sql/covid_27_svod_old.sql', 'r').read()
    SQL_3 = open('func/parus/sql/covid_27_svod_2.sql'  , 'r').read()
    SQL_4 = open('func/parus/sql/covid_27_svod_3.sql'  , 'r').read()
    SQL_5 = open('func/parus/sql/covid_27_svod_4.sql'  , 'r').read()

    DF      = parus_sql(SQL_1)
    DF_OLD  = parus_sql(SQL_2)
    DF_2    = parus_sql(SQL_3)
    DF_3    = parus_sql(SQL_4)
    DF_4    = parus_sql(SQL_5)

    DATE =  datetime.datetime.now().strftime('%d_%m_%Y')

    DF_3 = DF_3.loc[~(DF_3['IDMO'].isnull())]

    for col in DF_3.columns[10:]:
        try:
            DF_3[col] = DF_3[col].str.replace(',','.').astype(float)
        except:
            pass

    DF_4 = DF_4.loc[~(DF_4['IDMO'].isnull())]

    for col in DF_4.columns[10:]:
        try:
            DF_4[col] = DF_4[col].str.replace(',','.').astype(float)
        except:
            pass


    NEW_NAME_1 = 'temp/' + DATE + '_1_27_COVID-19.xlsx'
    NEW_NAME_2 = 'temp/' + DATE + '_2_Результаты_исследований_материала_на_COVID-19.xlsx'
    NEW_NAME_3 = 'temp/' + DATE + '_3_Кратность_лаб_обсл_на_COVID-19.xlsx'
    NEW_NAME_4 = 'temp/' + DATE + '_4_Кратность_положительных_лаб_обсл_на_COVID-19.xlsx'

    shutil.copyfile('help/27_COVID_19_svod.xlsx',   NEW_NAME_1)
    shutil.copyfile('help/27_COVID_19_svod_2.xlsx', NEW_NAME_2)
    shutil.copyfile('help/27_COVID_19_svod_3.xlsx', NEW_NAME_3)
    shutil.copyfile('help/27_COVID_19_svod_4.xlsx', NEW_NAME_4)

    wb= openpyxl.load_workbook( NEW_NAME_1 )
    ws = wb['Для заполнения']
    
    rows = dataframe_to_rows(DF,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    ws = wb['Пред.отч']
    rows = dataframe_to_rows(DF_OLD,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( NEW_NAME_1 )


    wb= openpyxl.load_workbook( NEW_NAME_2 )
    ws = wb['свод']
    rows = dataframe_to_rows(DF_2,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( NEW_NAME_2 )

    wb= openpyxl.load_workbook( NEW_NAME_3)
    ws = wb['свод']
    rows = dataframe_to_rows(DF_3,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( NEW_NAME_3 )


    wb= openpyxl.load_workbook( NEW_NAME_4)
    ws = wb['свод']
    rows = dataframe_to_rows(DF_4,index=False, header=False)
    for r_idx, row in enumerate(rows,4):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( NEW_NAME_4 )

    return NEW_NAME_1 + ';' + NEW_NAME_2 + ';' + NEW_NAME_3 + ';' + NEW_NAME_4



    

