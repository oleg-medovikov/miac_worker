import time, datetime, shutil, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from base import parus_sql

def svod_52_covid_19():
    SQL_1 = open('func/parus/sql/covid_52_svod.sql', 'r').read()
    SQL_2 = open('func/parus/sql/covid_52_svod_old.sql', 'r').read()

    DF  = parus_sql( SQL_1 )
    OLD = parus_sql( SQL_2 )

    DATE = DF.at[0,'DAY']

    del DF  ['DAY']
    del DF  ['ORGANIZATION']
    del OLD ['DAY']
    del OLD ['ORGANIZATION']

    NEW_NAME_1 = 'temp/' + DATE +  '_52_COVID_19_pred.xlsx' 
    NEW_NAME_2 = 'temp/' + DATE +  '_52_COVID_19_osn.xlsx'

    shutil.copyfile ( 'help/52_COVID_19_pred.xlsx', NEW_NAME_1 )
    shutil.copyfile ( 'help/52_COVID_19_osn.xlsx',  NEW_NAME_2 )

    wb= openpyxl.load_workbook( NEW_NAME_1 )
    
    ws = wb['52 COVID']
    rows = dataframe_to_rows(DF,index=False, header=False)
    for r_idx, row in enumerate(rows,11):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = wb['Вчера']
    rows = dataframe_to_rows(OLD,index=False, header=False)
    for r_idx, row in enumerate(rows,11):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save( NEW_NAME_1 )

    wb= openpyxl.load_workbook( NEW_NAME_2 )

    ws = wb['52 COVID']
    rows = dataframe_to_rows(DF,index=False, header=False)
    for r_idx, row in enumerate(rows,11):  
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save( NEW_NAME_2 )

    return NEW_NAME_1 + ';' + NEW_NAME_2


